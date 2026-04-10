"""One-time script to extract categorization rules and categories from Excel."""

import json
import sys

import openpyxl


def extract(excel_path: str, output_path: str):
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Extract categories from "Статьи" sheet
    ws_cat = wb['Статьи']
    categories = []
    for row in ws_cat.iter_rows(min_row=2, max_row=ws_cat.max_row, values_only=True):
        vals = list(row)
        name = vals[0] if vals else None
        if not name:
            continue
        categories.append({
            'name': str(name).strip(),
            'group': str(vals[1]).strip() if len(vals) > 1 and vals[1] else '',
            'activity_type': str(vals[2]).strip() if len(vals) > 2 and vals[2] else '',
            'description': str(vals[3]).strip() if len(vals) > 3 and vals[3] else '',
            'category_group': str(vals[4]).strip() if len(vals) > 4 and vals[4] else '',
        })

    # Extract rules from "Правила для реестра" sheet
    ws_rules = wb['Правила для реестра']
    rules = []
    seen = set()
    for row in ws_rules.iter_rows(min_row=2, max_row=ws_rules.max_row, values_only=True):
        vals = list(row)
        key = vals[0] if vals else None
        keyword = vals[1] if len(vals) > 1 else None
        category = vals[2] if len(vals) > 2 else None
        if not key or not keyword or not category:
            continue

        key_str = str(key).strip()
        keyword_str = str(keyword).strip()
        category_str = str(category).strip()

        # Map Excel field names to our field names
        if 'Назначение' in key_str:
            field = 'payment_description'
        elif 'Контрагент' in key_str:
            field = 'counterparty'
        else:
            field = 'payment_description'  # default

        # Dedup
        dedup_key = (field, keyword_str.lower(), category_str)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        rules.append({
            'field': field,
            'pattern': keyword_str,
            'category': category_str,
        })

    data = {
        'categories': categories,
        'rules': rules,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Extracted {len(categories)} categories and {len(rules)} rules to {output_path}")


if __name__ == '__main__':
    excel_path = sys.argv[1] if len(sys.argv) > 1 else 'managerial_reports.xlsx'
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'kz_finance/categorizer/categories.json'
    extract(excel_path, output_path)
