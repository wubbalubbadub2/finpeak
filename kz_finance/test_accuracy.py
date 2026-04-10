"""Measure categorization accuracy against the Excel ground truth."""

from __future__ import annotations

import sys
from collections import Counter, defaultdict
from datetime import date
from decimal import Decimal

import openpyxl

from .parsers.base import Transaction


def load_ground_truth(excel_path: str) -> list[dict]:
    """Load ground truth transactions from the 'Реестр платежей' sheet.

    Returns list of dicts with: date, amount, category, description, wallet.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Реестр платежей']

    # Find the header row (contains "Дата" and "Статья")
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(c) if c else "" for c in row]
        if any('Дата' in v for v in vals) and any('Статья' in v for v in vals):
            header_row = row_idx
            break

    if not header_row:
        print("ERROR: Could not find header row in Реестр платежей", file=sys.stderr)
        return []

    # Parse column indices from header
    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value) if cell.value else "")

    date_col = next((i for i, h in enumerate(headers) if 'Дата' == h.strip()), None)
    amount_col = next((i for i, h in enumerate(headers) if 'Сумма, тенге' in h or 'Сумма' in h), None)
    category_col = next((i for i, h in enumerate(headers) if 'Статья' in h), None)
    desc_col = next((i for i, h in enumerate(headers) if 'Назначение' in h), None)
    wallet_col = next((i for i, h in enumerate(headers) if 'Кошелек' in h), None)

    if date_col is None or category_col is None:
        print(f"ERROR: Missing columns. date={date_col}, category={category_col}", file=sys.stderr)
        return []

    # Parse data rows
    ground_truth = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals or len(vals) <= max(date_col, category_col):
            continue

        dt = vals[date_col]
        category = vals[category_col]
        if not dt or not category:
            continue

        # Parse date
        if isinstance(dt, str):
            from .parsers.base import parse_kz_date
            dt = parse_kz_date(dt)
        elif hasattr(dt, 'date'):
            dt = dt.date() if callable(getattr(dt, 'date')) else dt.date
        elif not isinstance(dt, date):
            continue

        # Parse amount
        amount = vals[amount_col] if amount_col is not None else None
        if amount is not None:
            try:
                amount = Decimal(str(amount))
            except Exception:
                amount = None

        desc = str(vals[desc_col] or "") if desc_col is not None else ""
        wallet = str(vals[wallet_col] or "") if wallet_col is not None else ""

        ground_truth.append({
            'date': dt,
            'amount': amount,
            'category': str(category).strip(),
            'description': desc[:50],
            'wallet': wallet.strip(),
        })

    return ground_truth


def measure_accuracy(
    parsed_transactions: list[Transaction],
    ground_truth: list[dict],
    wallet_filter: str = "",
) -> dict:
    """Compare parsed transactions against ground truth.

    Match by (date, abs_amount). Report categorization accuracy.
    """
    # Filter ground truth by wallet if specified
    gt = ground_truth
    if wallet_filter:
        gt = [g for g in gt if wallet_filter.lower() in g['wallet'].lower()]

    # Build lookup: (date, abs_amount) -> list of GT records
    gt_lookup = defaultdict(list)
    for g in gt:
        if g['amount'] is not None:
            key = (g['date'], abs(g['amount']))
            gt_lookup[key].append(g)

    # Match parsed transactions
    matched = 0
    category_correct = 0
    mismatches = []
    unmatched_parsed = []

    for tx in parsed_transactions:
        amount = abs(tx.signed_amount)
        key = (tx.date, amount)
        candidates = gt_lookup.get(key, [])

        if candidates:
            matched += 1
            gt_rec = candidates.pop(0)  # consume the first match
            if not gt_lookup[key]:
                del gt_lookup[key]

            if tx.category == gt_rec['category']:
                category_correct += 1
            elif tx.category:
                mismatches.append({
                    'date': tx.date,
                    'amount': tx.signed_amount,
                    'predicted': tx.category,
                    'actual': gt_rec['category'],
                    'description': tx.payment_description[:50],
                })
        else:
            unmatched_parsed.append(tx)

    total_parsed = len(parsed_transactions)
    parse_rate = matched / total_parsed if total_parsed > 0 else 0
    cat_accuracy = category_correct / matched if matched > 0 else 0

    return {
        'total_parsed': total_parsed,
        'total_ground_truth': len(gt),
        'matched': matched,
        'parse_rate': parse_rate,
        'category_correct': category_correct,
        'category_accuracy': cat_accuracy,
        'mismatches': mismatches,
        'unmatched_parsed': len(unmatched_parsed),
    }


def print_accuracy_report(results: dict):
    """Print a formatted accuracy report."""
    print(f"\n{'=' * 60}")
    print(f"  ACCURACY REPORT")
    print(f"{'=' * 60}")
    print(f"  Parsed transactions:    {results['total_parsed']}")
    print(f"  Ground truth entries:   {results['total_ground_truth']}")
    print(f"  Matched:                {results['matched']}")
    print(f"  Parse rate:             {results['parse_rate']*100:.1f}%")
    print(f"  Category correct:       {results['category_correct']}")
    print(f"  Category accuracy:      {results['category_accuracy']*100:.1f}%")
    print(f"  Unmatched (parsed):     {results['unmatched_parsed']}")
    print(f"{'=' * 60}")

    if results['mismatches']:
        print(f"\n  MISMATCHES ({len(results['mismatches'])}):")
        for m in results['mismatches'][:20]:
            print(f"    {m['date']} | {m['amount']:>12} | predicted: {m['predicted'][:30]}")
            print(f"    {'':>14} | {'':>12} | actual:    {m['actual'][:30]}")
            print(f"    {'':>14} | desc: {m['description']}")
            print()
